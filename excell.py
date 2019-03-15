from openpyxl import load_workbook
from openpyxl import Workbook

#Thư viện open file
from tkinter import filedialog
from tkinter import *
root = Tk()
root.withdraw()
root.filename = filedialog.askopenfilename(initialdir = "/",title = "Chọn file",filetypes = (("Excell 2010","*.xlsx"),("all files","*.*")))

wb = Workbook()
wb = load_workbook(root.filename)
ws = wb.active
# add a simple formula
wb["D4"] = "=SUM(1,1)"
root.destroy()